
#!/usr/bin/env python
# -*- coding: utf-8 -*-

# Contact email: dimitris.rizopoulos@inlecomsystems.com

import os
import numpy as np
import pickle5 as pickle
from openpyxl import load_workbook

script_dir = os.path.dirname(__file__)


def read_pickle(file_path):
    result = []

    if os.path.isfile(file_path):
        with open(file_path, 'rb') as handle:
            result = pickle.load(handle)

    return result

def read_xlsx(file_path):
    workbook_1 = load_workbook(file_path)
    worksheets_1 = workbook_1.sheetnames

    # N is the number of rows in the input sheet
    N_1 = 0
    M_1 = 0
    for row in workbook_1[worksheets_1[0]]:
        N_1 = N_1 + 1
        if (N_1 == 1):
            for column in row:
                M_1 = M_1 + 1

    result = np.empty((N_1 - 1, M_1), dtype=object)
    for row in range(1, N_1):
        for column in range(M_1):
            row_1 = row - 1
            result[row_1][column] = workbook_1[worksheets_1[0]].cell(row=row + 1, column=column + 1).value

    return result


def dump_xlsx_to_pickle(xlsx_path, pickle_path):
    result = read_xlsx(xlsx_path)

    with open(pickle_path, 'wb') as handle:
        pickle.dump(result, handle, protocol=pickle.HIGHEST_PROTOCOL)

def dump_table_to_pickle(table, pickle_path):
    with open(pickle_path, 'wb') as handle:
        pickle.dump(table, handle, protocol=pickle.HIGHEST_PROTOCOL)

def read_file(base_dir, input_file, create_pickle=True):
    file_name = os.path.splitext(input_file)[0]
    pickle_file_path = os.path.abspath(os.path.join(
        base_dir,
        file_name + '.pickle'))

    xlsx_file_path = os.path.abspath(os.path.join(
        base_dir,
        file_name + '.xlsx'))

    if os.path.exists(pickle_file_path):
        result = read_pickle(pickle_file_path)
    elif os.path.exists(xlsx_file_path):
        result = read_xlsx(xlsx_file_path)
        if create_pickle:
            dump_xlsx_to_pickle(xlsx_file_path, pickle_file_path)
    else:
        raise Exception("File:{} does not exists".format(file_name))

    return result
